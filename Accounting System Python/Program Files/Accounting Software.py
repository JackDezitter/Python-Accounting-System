
#This system places inputted tax data into a templated spreadsheet
#Converts the spreadsheet do a dataframe using pandas, Sorts the data chronologically
#Calculates monthly expenses, monthly income, annual income, annual expense and annual profit
#Places the sorted data back into the original file and overwrites the old data
#Storing a sorted file of the data

##Libraries##

#Used to create and use the inferface
from guizero import App, Box, Text, PushButton, TextBox, Window, error, info, warn, yesno
#Openpyxl manipulates excel spreadsheets
from openpyxl import Workbook, load_workbook
#Allows the system to convert inputted strings into dates to be sorted
import datetime
#Sorts the data and calculate 
import pandas as pd
#Manages file locations
import os




global startYear
global endYear
global taxDates
global taxDatesFile
global filePath
global wb

##Code Functions##

#Checks if file exists to edit and calls it
def editCheck(startYearText):

    global startYear
    global endYear
    global taxDates
    global taxDatesFile
    global filePath
    global wb
    global ws
    
    #Takes the starting year of the account, calculates the end year
    #Creates name for file
    startYear = int(startYearText.value)
    endYear = startYear + 1
    taxDates = "6Apr"+str(startYear)+"-5Apr"+str(endYear)
    #Creates file name
    taxDatesFile = taxDates+".xlsx"
    
    #Checks if file exists in the correct folder
    #Retrieves the programs current directory path
    currentDirect = os.path.dirname(os.path.realpath(__file__))
    #Steps back once in the path and changes path to correct folder
    accountsDirect = os.path.dirname(currentDirect) + "\Accounts"
    #Creates path where the file would exist
    filePath = accountsDirect + "\\" + taxDatesFile

    #Searches if file with year title exist
    #Returns False if not
    #Returns Truth if it does
    check = os.path.isfile(str(filePath))
    #If file exists 
    if check == True:
        selectWindow.hide()
        wb = load_workbook(filePath)
        save()
        ws = wb.active
        rowCount()
        
        
    else:
        error("ERROR", text = "This file does not exist")
        
        
    
        

#Checks if file exists to create file and creates it
def newCheck(startYearText):

    global startYear
    global endYear
    global taxDates
    global taxDatesFile
    global filePath
    global wb
    global ws

    
    #Takes the starting year of the account, calculates the end year
    #Creates name for file
    startYear = int(startYearText.value)
    endYear = startYear + 1
    taxDates = "6Apr"+str(startYear)+"-5Apr"+str(endYear)
    #Creates file name
    taxDatesFile = taxDates+".xlsx"
    
    #Checks if file exists in the correct folder
    #Retrieves the programs current directory path
    currentDirect = os.path.dirname(os.path.realpath(__file__))
    #Steps back once in the path and changes path to correct folder
    accountsDirect = os.path.dirname(currentDirect) + "\Accounts"
    #Creates path where the file would exist
    filePath = accountsDirect + "\\" + taxDatesFile

    #Searches if file with year title exist
    #Returns False if not
    #Returns Truth if it does
    check = os.path.isfile(str(filePath))

    #If file doesn't exist 
    if check == False:
        selectWindow.hide()

        #Loads Template for new file and Titles it the correct dates
        wb = load_workbook('Template.xlsx')
        wb.template = False
        ws = wb.active

        #Fills in date of the spreadsheet for which dates the current sheet is for
        ws['B2'] = taxDates
        save()
        
        rowCount()

    else:
        error("ERROR", text = "This file already exist")

#Save file of excel spreadsheet shortcut for programming
def save():
    wb.save(filePath)

#Checks number of rows in spreadsheet to input data
def rowCount():
    global inputCell
    
    #Creates dataframe of current excel file from column A-F
    tax = pd.read_excel(filePath, usecols = "A:F")

    #Removes rows created by month titles outside of bounds
    tax = tax.dropna(how = 'all')
    #Checks the number of rows in the file so data is inputted on the next line
    inputCell = tax.shape[0]+2

    inputWindowCreate()

def inputWindowCreate():
    global inputWindow
    global inputBox
    
    inputWindow = Window(selectWindow, title = "Accounting Software Data Input", height = 400, width =900) 

    spaceBox = Box(inputWindow, align  = "top", width = "fill")
    space = Text(spaceBox, text = "   ") # This creates a space


    spaceBox = Box(inputWindow, align  = "top", width = "fill")
    space = Text (spaceBox, text = "   ") # This creates a space

    inputTitle = Text(inputWindow, align = "top", text = "Please Input your data below")

    spaceBox = Box(inputWindow, align  = "top", width = "fill")
    space = Text (spaceBox, text = "   ") # This creates a space

    inputBox = Box(inputWindow, align = "top", width = "fill")

    expense()
    income()

    spaceBox = Box(inputWindow, align  = "top", width = "fill")
    space = Text (spaceBox, text = "   ") # This creates a space

    sortButton = PushButton(inputWindow, align = "top", text = "Sort Data", command = sort)

    menuBox = Box(inputWindow, align = "bottom", width = "fill")
    menuButton = PushButton(menuBox, text = "Return to Menu", align = "right", command = menuConfirm)

def menuConfirm():
    
    confirm = yesno("Menu","Return to Main Menu?")
    if confirm == True:
        inputWindow.destroy()
        selectWindow.show()
        
    else:
        inputWindow.show()
    



#Submit Button on Expenses and places expense data
def expenseSubmit(date, category, description, expense, paymentMethod):

    global inputCell
    
    date = date.value
    category = category.value
    description = description.value
    expense = expense.value
    paymentMethod = paymentMethod.value

    #Validates that the date is in the correct format using length check
    if len(date)!= 10:
        error("Input Error", "Please input date in form DD/MM/YYYY")
    else:

        #Makes the data be inputted into the next row
        dateColumn = "A"+str(inputCell)
        categoryColumn = "B"+str(inputCell)
        descriptionColumn = "C"+str(inputCell)
        expenseColumn = "E"+str(inputCell)
        paymentMethodColumn = "F"+str(inputCell)

        #Inputs Data into row for expense
        ws[dateColumn] = date
        ws[categoryColumn] = category
        ws[descriptionColumn] = description
        ws[expenseColumn] = expense
        ws[paymentMethodColumn] = paymentMethod
        save()

        inputCell = inputCell + 1

        info("Data Entered", "Expense stored")
    

#Clears text boxes in expenses 
def expenseClear(date, category, description, expense, paymentMethod):

    date.clear()
    category.clear()
    description.clear()
    expense.clear()
    paymentMethod.clear()


#Collects expense Data
def expense():

    global inputCell

    #Collects Data for expense
    #Creates Box to keep all expese inputs together
    expenseDataBox = Box(inputBox, align = "left", border = True)
    
    #Creates Title
    expenseTitle = Text(expenseDataBox, align = "top", text = "Expenses:")

    #Date Input
    dateBox = Box(expenseDataBox, align = "top", width = "fill", border = True)
    dateText = Text(dateBox, text = "Date of Purchase (DD/MM/YYYY):  ", align = "left")
    date = TextBox(dateBox, align = "left", width = 20)

    #Category Input
    categoryBox = Box(expenseDataBox, align = "top", width = "fill", border = True)
    categoryText = Text(categoryBox, text = "Category of Purchase:  ", align = "left")
    category = TextBox(categoryBox, align = "left", width = 20)
    
    #Description Input
    descriptionBox = Box(expenseDataBox, align = "top", width = "fill", border = True)
    descriptionText = Text(descriptionBox, text = "Description of Purchase:  ", align = "left")
    description = TextBox(descriptionBox, align = "left", width = 20)

    #Expense Input
    expenseBox = Box(expenseDataBox, align = "top", width = "fill", border = True)
    expenseText = Text(expenseBox, text = "Expense Paid:  ", align = "left")
    expense = TextBox(expenseBox, align = "left", width = 20)

    #Payment Method Input
    paymentMethodBox = Box(expenseDataBox, align = "top", width = "fill", border = True)
    paymentMethodText = Text(paymentMethodBox, text = "Payment Method:  ", align = "left")
    paymentMethod = TextBox(paymentMethodBox, align = "left", width = 20)


    #Creates Button box so once the user has inputted the data
    #They can submit it to be placed into the spreadsheet
    #Then clear the Textboxes for a new entry
    buttonBox = Box(expenseDataBox, align = "top")

    
    #Submit Button
    submitButtonBox = Box(buttonBox,align = "left")
    submitButton = PushButton(submitButtonBox, align = "right", text = "Submit", command = expenseSubmit,
                              args = [date, category, description, expense, paymentMethod])

    #Clear Button
    clearButtonBox = Box(buttonBox, align = "right")
    clearButton = PushButton(clearButtonBox, align = "left", text = "Clear", command = expenseClear,
                             args = [date, category, description, expense, paymentMethod])


#Submit Button on Income and places Income data
def incomeSubmit(date, description, income, paymentMethod):

    global inputCell

    #Takes inputed data as variables to be placed
    date = date.value
    category = "Income"
    description = description.value
    income = income.value
    paymentMethod = paymentMethod.value

    #Validates that the date is in the correct format using length check
    if len(date)!= 10:
        error("Input Error", "Please input date in the form DD/MM/YYYY")
    else:
        #Makes the date be inputted into the next row
        dateColumn = "A"+str(inputCell)
        categoryColumn = "B"+str(inputCell)
        descriptionColumn = "C"+str(inputCell)
        incomeColumn = "D"+str(inputCell)
        paymentMethodColumn = "F"+str(inputCell)

        #Inputs Data into row 
        ws[dateColumn] = date
        ws[categoryColumn] = "Income"
        ws[descriptionColumn] = description
        ws[incomeColumn] = income
        ws[paymentMethodColumn] = paymentMethod
        save()

        inputCell = inputCell + 1

        info("Data Entered", "Income stored")


#Clears income text boxes for new data entry
def incomeClear(date, description, income, paymentMethod):

    date.clear()
    description.clear()
    income.clear()
    paymentMethod.clear()

#Collects income data
def income():

    global inputCell
    
    #Creates box and title for income data
    incomeDataBox = Box(inputBox, align = "right", border = True)
    incomeTitle = Text(incomeDataBox, align = "top", text = "Income:")
    
    #Date Input
    dateBox = Box(incomeDataBox, align = "top", width = "fill", border = True)
    dateText = Text(dateBox, text = "Date of Purchase (DD/MM/YYYY):  ", align = "left")
    date = TextBox(dateBox, align = "left", width = 20)

    #Description Input
    descriptionBox = Box(incomeDataBox, align = "top", width = "fill", border = True)
    descriptionText = Text(descriptionBox, text = "Description of Income:  ", align = "left")
    description = TextBox(descriptionBox, align = "left", width = 20)

    #Income Amount Input
    incomeBox = Box(incomeDataBox, align = "top", width = "fill", border = True)
    incomeText = Text(incomeBox, text = "Income Amount:  ", align = "left")
    income = TextBox(incomeBox, align = "left", width = 20)

    #Payment Method Input
    paymentMethodBox = Box(incomeDataBox, align = "top", width = "fill", border = True)
    paymentMethodText = Text(paymentMethodBox, text = "Payment Method:  ", align = "left")
    paymentMethod = TextBox(paymentMethodBox, align = "left", width = 20)

    spaceBox = Box(incomeDataBox, align  = "top", width = "fill")
    space = Text(spaceBox, text = "   ") # This creates a space
    
    #Creates Button box so once the user has inputted the data
    #They can submit it to be placed into the spreadsheet
    #Then clear the Textboxes for a new entry
    buttonBox = Box(incomeDataBox, align = "top")

    submitButtonBox = Box(buttonBox,align = "left")
    submitButton = PushButton(submitButtonBox, align = "right", text = "Submit", command = incomeSubmit,
                              args = [date, description, income, paymentMethod])
    
    clearButtonBox = Box(buttonBox, align = "right")
    clearButton = PushButton(clearButtonBox, align = "left", text = "Clear", command = incomeClear,
                             args = [date, description, income, paymentMethod])


def sort():

    ##Sorting System##

    #Imports excel into panda and sorts by date
    #Skips first 4 rows so Row 5 = headers
    #Only takes columns A-F as they hold the data entries
    tax = pd.read_excel(filePath, skiprows=4, usecols = "A:F" )

    #Converts strings in date column to dates
    tax['Date'] = pd.to_datetime(tax.Date,dayfirst = True)
    #Removes rows of no data
    tax = tax.dropna(how = 'all')
    #Fills empty spaces in the data with 0.0 (Income and expense data)
    tax = tax.fillna(0)

    #Sorts the data by date
    tax = tax.sort_values(['Date'], ascending=True,)#Sorts date in ascending order

    formula(tax)

    #Puts date format and floats back to string for insertion as only strings can be iterated
    tax['Date'] = tax['Date'].dt.strftime('%d/%m/%Y')
    tax['Income'] = tax['Income'].astype(str)
    tax['Amount'] = tax['Amount'].astype(str)


    #Creates new file from panda database to an excel file
    taxDatesPanda = 'PD'+taxDatesFile
    tax.to_excel(taxDatesPanda, 'Sheet1',header = False, index = False)

    wb2 = load_workbook(taxDatesPanda)
    ws2 = wb2['Sheet1']

    #Informs the user the program is running
    warn("Sorting","Please wait, your Data is being sorted, please keep this window open")

    #Transfers each cell that has a value from the sorted data to the original template file
    #Sorted data starts on row 1
    organisedCell = 1
    #Unsorted data starts to be ovrwriten by sorted data at row 6 as headings end in row 5
    placedCell = 6
    #Checks which columns in the sorted data has values and if there is a value it will assign it
    #to the corresponding cell in the unsorted data(replacing the old data)
    for column in ws2.values:
        for value in column[0]:
            ws['A' + str(placedCell)] = ws2['A' +str(organisedCell)].value
            #Increments counters by 1 so it takes the next cell in the column
            organisedCell = organisedCell + 1
            placedCell = placedCell + 1
            #Would save the data after each entry
            save()
        #counters are reset
        #Process repeats of each column 
        organisedCell = 1
        placedCell = 6
        for value in column[1]:
            ws['B' + str(placedCell)] = ws2['B' + str(organisedCell)].value
            #Increments counters by 1 so it takes the next cell in the column
            organisedCell = organisedCell + 1
            placedCell = placedCell + 1
            #Would save the data after each entry
            save()

        #counters are reset
        #Process repeats of each column 
        organisedCell = 1
        placedCell = 6
        for value in column[2]:
            ws['C' + str(placedCell)] = ws2['C' + str(organisedCell)].value
            #Increments counters by 1 so it takes the next cell in the column
            organisedCell = organisedCell + 1
            placedCell = placedCell + 1
            #Would save the data after each entry
            save()

        #counters are reset
        #Process repeats of each column 
        organisedCell = 1
        placedCell = 6
        for value in column[3]:
            ws['D' + str(placedCell)] = ws2['D' + str(organisedCell)].value
            #Increments counters by 1 so it takes the next cell in the column
            organisedCell = organisedCell + 1
            placedCell = placedCell + 1
            #Would save the data after each entry
            save()

        #counters are reset
        #Process repeats of each column
        organisedCell = 1
        placedCell = 6
        for value in column[4]:
            ws['E' + str(placedCell)] = ws2['E' + str(organisedCell)].value
            #Increments counters by 1 so it takes the next cell in the column
            organisedCell = organisedCell + 1
            placedCell = placedCell + 1
            #Would save the data after each entry
            save()

        #counters are reset
        #Process repeats of each column
        organisedCell = 1
        placedCell = 6
        for value in column[5]:
            ws['F' + str(placedCell)] = ws2['F' + str(organisedCell)].value
            #Increments counters by 1 so it takes the next cell in the column
            organisedCell = organisedCell + 1
            placedCell = placedCell + 1
            #Would save the data after each entry
            save()



        #Deletes Panda Database file
        os.remove(taxDatesPanda)

        exitProgram = yesno("Sort Complete", "Would you like to exit this program?")
        if exitProgram == True:
            selectWindow.destroy()
            print("l")
            
        else:
            inputWindow.show()



    

#Calculates monthly and annual data
def formula(tax):
    #Establishes global variables for each total
    global totalIncome
    global totalExpense
    global balance
    global janIncome
    global janExpense
    global febIncome
    global febExpense
    global marIncome
    global marExpense
    global aprIncome
    global aprExpense
    global mayIncome
    global mayExpense
    global junIncome
    global junExpense
    global julIncome
    global julExpense
    global augIncome
    global augExpense
    global sepIncome
    global sepExpense
    global octoIncome
    global octoExpense
    global novIncome
    global novExpense
    global decIncome
    global decExpense


    
    
    
    #creates temporary dataframe so main data is not altered
    tempData = tax
    #Calculates total income and expense then overall profit
    totalIncome = tempData['Income'].sum()
    totalExpense = tempData['Amount'].sum()
    balance = totalIncome - totalExpense

    #Creates column with the month of the entry
    tempData['Month'] = pd.DatetimeIndex(tempData['Date']).month
    #Groups data by month
    tempData = tempData.groupby(['Month']).sum()
    # Calculates sum of income and expense for each month

    #January
    #Checks if there is a value for month
    janCheck = 1 in tempData.index.values
    #If returns true, new dataframe created then month totals are assigned, if not they keep default of 0.0
    if janCheck == True:
        jan = tempData.loc[1]
        janIncome = jan['Income'].sum()
        janExpense = jan['Amount'].sum()
    #If value is not in the index then the totals for that month is 0.0
    elif janCheck == False:
        janIncome = 0.0
        janExpense = 0.0

    #February
    febCheck = 2 in tempData.index.values
    if febCheck == True:
        feb = tempData.loc[2]
        febIncome = feb['Income'].sum()
        febExpense = feb['Amount'].sum()

    elif febCheck == False:
        febIncome = 0.0
        febExpense = 0.0
    
    
    #March
    marCheck = 3 in tempData.index.values
    if marCheck == True:
        mar = tempData.loc[3]
        marIncome = mar['Income'].sum()
        marExpense = mar['Amount'].sum()

    elif marCheck == False:
        marIncome = 0.0
        marExpense = 0.0

    #April
    aprCheck = 4 in tempData.index.values
    if aprCheck == True:
        apr = tempData.loc[4]
        aprIncome = apr['Income'].sum()
        aprExpense = apr['Amount'].sum()

    elif aprCheck == False:
        aprIncome = 0.0
        aprExpense = 0.0

    #May
    mayCheck = 5 in tempData.index.values
    if mayCheck == True:
        may = tempData.loc[5]
        mayIncome = may['Income'].sum()
        mayExpense = may['Amount'].sum()

    elif mayCheck == False:
        mayIncome = 0.0
        mayExpense = 0.0

    #June
    junCheck = 6 in tempData.index.values
    if junCheck == True:
        jun = tempData.loc[6]
        junIncome = jun['Income'].sum()
        junExpense = jun['Amount'].sum()

    elif junCheck == False:
        junIncome = 0.0
        junExpense = 0.0

    #July
    julCheck = 7 in tempData.index.values
    if julCheck == True:
        jul = tempData.loc[7]
        julIncome = jul['Income'].sum()
        julExpense = jul['Amount'].sum()

    elif julCheck == False:
        julIncome = 0.0
        julExpense = 0.0

    #August
    augCheck = 8 in tempData.index.values
    if augCheck == True:
        aug = tempData.loc[8]
        augIncome = aug['Income'].sum()
        augExpense = aug['Amount'].sum()

    elif augCheck == False:
        augIncome = 0.0
        augExpense = 0.0
        
    #September
    sepCheck = 9 in tempData.index.values
    if sepCheck == True:
        sep = tempData.loc[9]
        sepIncome = sep['Income'].sum()
        sepExpense = sep['Amount'].sum()

    elif sepCheck == False:
        sepIncome = 0.0
        sepExpense = 0.0

    #October
    #Cannot use oct as is a process in python
    octoCheck = 10 in tempData.index.values
    if octoCheck == True:
        octo = tempData.loc[10]
        octoIncome = octo['Income'].sum()
        octoExpense = octo['Amount'].sum()

    elif octoCheck == False:
        octoIncome = 0.0
        octoExpense = 0.0

    #November
    novCheck = 11 in tempData.index.values
    if novCheck == True:
        nov = tempData.loc[11]
        novIncome = nov['Income'].sum()
        novExpense = nov['Amount'].sum()

    elif novCheck == False:
        novIncome = 0.0
        novExpense = 0.0

    #Decemeber
    decCheck = 12 in tempData.index.values
    if decCheck == True:
        dec = tempData.loc[12]
        decIncome = dec['Income'].sum()
        decExpense = dec['Amount'].sum()

    elif decCheck == False:
        decIncome = 0.0
        decExpense = 0.0

        monthData()
    
    
    
    
#Placed month and annual data into the spreadsheet
def monthData():
    
    #Places Annual Data
    ws['D3'] = "£" + str(totalIncome)
    ws['E3'] = "£" + str(totalExpense)
    ws['G3'] = "£" + str(balance)

    #Starting row for month data
    monthTotalCell = 6

    #Places total income for each month into the column J
    #April
    ws['J' + str(monthTotalCell)] = "£" + str(aprIncome)
    monthTotalCell = monthTotalCell + 1
    #May
    ws['J' + str(monthTotalCell)] = "£" + str(mayIncome)
    monthTotalCell = monthTotalCell + 1
    #June
    ws['J' + str(monthTotalCell)] = "£" + str(junIncome)
    monthTotalCell = monthTotalCell + 1
    #July
    ws['J' + str(monthTotalCell)] = "£" + str(julIncome)
    monthTotalCell = monthTotalCell + 1
    #August
    ws['J' + str(monthTotalCell)] = "£" + str(augIncome)
    monthTotalCell = monthTotalCell + 1
    #September
    ws['J' + str(monthTotalCell)] = "£" + str(sepIncome)
    monthTotalCell = monthTotalCell + 1
    #October
    ws['J' + str(monthTotalCell)] = "£" + str(octoIncome)
    monthTotalCell = monthTotalCell + 1
    #November
    ws['J' + str(monthTotalCell)] = "£" + str(novIncome)
    monthTotalCell = monthTotalCell + 1
    #December
    ws['J' + str(monthTotalCell)] = "£" + str(decIncome)
    monthTotalCell = monthTotalCell + 1
    #January
    ws['J' + str(monthTotalCell)] = "£" + str(janIncome)
    monthTotalCell = monthTotalCell + 1
    #February
    ws['J' + str(monthTotalCell)] = "£" + str(febIncome)
    monthTotalCell = monthTotalCell + 1
    #March
    ws['J' + str(monthTotalCell)] = "£" + str(marIncome)
    monthTotalCell = monthTotalCell + 1

    #Saves spreadsheet
    save()

    #Places total expenses for each month in column K
    #Resets starting row 
    monthTotalCell = 6
    #Places total expenses for each month in column K
    #April
    ws['K' + str(monthTotalCell)] = "£" + str(aprExpense)
    monthTotalCell = monthTotalCell + 1
    #May
    ws['K' + str(monthTotalCell)] = "£" + str(mayExpense)
    monthTotalCell = monthTotalCell + 1
    #June
    ws['K' + str(monthTotalCell)] = "£" + str(junExpense)
    monthTotalCell = monthTotalCell + 1
    #July
    ws['K' + str(monthTotalCell)] = "£" + str(julExpense)
    monthTotalCell = monthTotalCell + 1
    #August
    ws['K' + str(monthTotalCell)] = "£" + str(augExpense)
    monthTotalCell = monthTotalCell + 1
    #September
    ws['K' + str(monthTotalCell)] = "£" + str(sepExpense)
    monthTotalCell = monthTotalCell + 1
    #October
    ws['K' + str(monthTotalCell)] = "£" + str(octoExpense)
    monthTotalCell = monthTotalCell + 1
    #November
    ws['K' + str(monthTotalCell)] = "£" + str(novExpense)
    monthTotalCell = monthTotalCell + 1
    #December
    ws['K' + str(monthTotalCell)] = "£" + str(decExpense)
    monthTotalCell = monthTotalCell + 1
    #January
    ws['K' + str(monthTotalCell)] = "£" + str(janExpense)
    monthTotalCell = monthTotalCell + 1
    #February
    ws['K' + str(monthTotalCell)] = "£" + str(febExpense)
    monthTotalCell = monthTotalCell + 1
    #March
    ws['K' + str(monthTotalCell)] = "£" + str(marExpense)
    monthTotalCell = monthTotalCell + 1

    #Saves spreadsheet
    save()
    

#Window to input start year and whether the user wants to edit or create
#a new file
#Creates window

selectWindow = App(title="Accounting Software", height = 270)

#Creates Title
titleBox = Box(selectWindow, align = "top", width = "fill")
title = Text(titleBox, text="Welcome Lora Dezitter")

#Creates a space to aid aesthetics
spaceBox = Box(selectWindow, align  = "top", width = "fill")
space = Text (spaceBox, text = "   ") # This creates a space

#Input the starting year for the file
startYearBox = Box(selectWindow, layout = "grid")
Text(startYearBox, text = "Please input the starting year of the account:", grid = [5,0])
startYearText = TextBox(startYearBox, grid = [6,0])
#Assigns inputted value to variable


#Creates a space to aid aesthetics
spaceBox = Box(selectWindow, align  = "top", width = "fill")
space = Text (spaceBox, text = "   ") # This creates a space

#Text to ask the user to select between creating or editing an account
captionBox = Box(selectWindow, align = "top", width = "fill")
caption = Text(captionBox, text = "Would you like to create a new account or edit an existing one?")

spaceBox = Box(selectWindow, align  = "top", width = "fill")
space = Text (spaceBox, text = "   ") # This creates a space


buttonBox = Box(selectWindow, align = "top", width = "fill")
editBox = Box(buttonBox, align="right",width = "fill")
editButton = PushButton(editBox, align = "top", command = editCheck, args = [startYearText],
                        text = "Press edit")

newBox = Box(buttonBox, align="left",width = "fill")
newButton = PushButton(newBox, align = "top", command = newCheck, args = [startYearText],
                       text = "Press new")


selectWindow.display()
